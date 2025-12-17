#!/usr/bin/env python3
"""
Find top-level expenditure categories in 2023 CE table (like Food, Housing, Apparel, etc.)
"""

import openpyxl
from pathlib import Path

file_path = Path("data/raw/ce_quintile_2023.xlsx")

wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb.active

print("Looking for major expenditure categories (not SE/RSE subcategories)...")
print("=" * 120)

# Skip SE/RSE/detail rows - look for major categories only
expenditure_items = []

for row_idx in range(50, min(sheet.max_row + 1, 200)):  # Start after demographics
    col_a = sheet.cell(row_idx, 1).value
    
    if not col_a or not isinstance(col_a, str):
        continue
    
    item = col_a.strip()
    
    # Skip SE, RSE, and very detailed subcategories
    if item in ['SE', 'RSE', 'Mean', 'Share']:
        continue
    if item.startswith('SE') or item.startswith('RSE'):
        continue
    
    # Look for Mean row in next few rows
    found_mean_share = False
    for offset in range(1, 4):
        if row_idx + offset >= sheet.max_row:
            break
        
        check_val = sheet.cell(row_idx + offset, 1).value
        if check_val and str(check_val).strip() == "Mean":
            # Check if next row is Share
            if row_idx + offset + 1 < sheet.max_row:
                next_val = sheet.cell(row_idx + offset + 1, 1).value
                if next_val and str(next_val).strip() == "Share":
                    found_mean_share = True
                    
                    # Get share values
                    share_row = row_idx + offset + 1
                    shares = []
                    for col in range(2, 8):
                        val = sheet.cell(share_row, col).value
                        if isinstance(val, (int, float)):
                            shares.append(val)
                    
                    if len(shares) >= 5 and max(shares) > 1.0:  # Significant shares (>1%)
                        expenditure_items.append({
                            'label': item,
                            'row': row_idx,
                            'shares': shares[:5],
                            'total': sum(shares[:5])
                        })
                        print(f"Row {row_idx:3d}: {item:60s} | Shares: Q1={shares[0]:5.1f}% Q2={shares[1]:5.1f}% ... | Total={sum(shares[:5]):5.1f}%")
                    break

print("=" * 120)
print(f"\nFound {len(expenditure_items)} major expenditure categories")

# Sort by total to see which are most significant
expenditure_items.sort(key=lambda x: x['total'], reverse=True)

print("\nTop categories by total share:")
for item in expenditure_items[:15]:
    print(f"  {item['total']:6.1f}% - {item['label']}")
