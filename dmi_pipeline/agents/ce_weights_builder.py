"""
CE Weights Builder - Extract and map CE table shares to CPI category weights.

This module implements weight extraction from BLS CE published tables
and maps CE items to CPI categories using the pinned mapping artifact.
"""

import json
from pathlib import Path
from typing import Dict, List, Tuple, Optional

import openpyxl
import pandas as pd
import numpy as np


def extract_weights_from_ce_table(
    file_path: Path,
    table_type: str,
    mapping_path: Path
) -> pd.DataFrame:
    """
    Extract weights from CE published table and map to CPI categories.
    
    Args:
        file_path: Path to validated CE XLSX file
        table_type: "quintile" or "decile"
        mapping_path: Path to ce_table_to_cpi_mapping_v0_1.json
    
    Returns:
        DataFrame with columns [group_id, category_id, weight, raw_share, ce_item]
    
    Raises:
        ValueError: If extraction fails
    """
    # Load mapping
    with open(mapping_path) as f:
        mapping_data = json.load(f)
    
    # Create CE item → CPI category mapping
    ce_to_cpi = {}
    excluded_items = set()
    
    for row in mapping_data['rows']:
        ce_label = row['ce_item_label']
        if row['include_in_inflation_universe']:
            ce_to_cpi[ce_label] = row['cpi_category_id']
        else:
            excluded_items.add(ce_label)
    
    # Load workbook
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active
    
    # Extract shares for each CE item
    shares_data = _extract_share_rows(sheet, table_type)
    
    # Map to CPI categories and aggregate
    weights_df = _map_and_aggregate_weights(
        shares_data, ce_to_cpi, excluded_items, table_type
    )
    
    return weights_df


def _extract_share_rows(sheet, table_type: str) -> pd.DataFrame:
    """
    Extract 'Share' rows from CE table.
    
    Returns DataFrame with columns [ce_item, group1_share, group2_share, ...]
    """
    # Find header row with income groups
    header_row_idx = None
    group_col_indices = []
    
    for row_idx in range(1, min(50, sheet.max_row + 1)):
        row_vals = [sheet.cell(row_idx, col).value for col in range(1, 20)]
        
        # Look for income group indicators in header
        group_keywords = ["lowest", "first", "second", "highest"]
        if any(kw in str(v).lower() if v else False for v in row_vals for kw in group_keywords):
            header_row_idx = row_idx
            # Find group column indices
            for col_idx, val in enumerate(row_vals, start=1):
                if val and isinstance(val, str):
                    val_lower = val.lower()
                    if any(kw in val_lower for kw in ["lowest", "first", "second", "third", 
                                                       "fourth", "fifth", "sixth", "seventh",
                                                       "eighth", "ninth", "tenth", "highest"]):
                        group_col_indices.append(col_idx)
            break
    
    if not header_row_idx or not group_col_indices:
        raise ValueError("Could not locate income group header row")
    
    # Expected number of groups
    expected_groups = {"quintile": 5, "decile": 10}[table_type]
    if len(group_col_indices) < expected_groups:
        raise ValueError(
            f"Found {len(group_col_indices)} group columns, expected at least {expected_groups}"
        )
    
    # Use first N groups (exclude "All consumer units" if present)
    group_col_indices = group_col_indices[:expected_groups]
    
    # Scan for CE item labels and their Share rows
    shares_records = []
    
    max_row = min(sheet.max_row, 600)  # Increased to capture all major categories (Transportation at row 399, Healthcare at 465, etc.)
    for row_idx in range(header_row_idx + 1, max_row + 1):
        col_a = sheet.cell(row_idx, 1).value
        
        if not col_a:
            continue
        
        # Check if this is a CE item label row (not "Mean" or "Share")
        if isinstance(col_a, str) and col_a.strip() not in ["Mean", "Share", "SE"]:
            ce_item = col_a.strip()
            
            # Look for Share row in next few rows
            for offset in range(1, 5):
                if row_idx + offset > max_row:
                    break
                
                check_row_label = sheet.cell(row_idx + offset, 1).value
                if check_row_label and str(check_row_label).strip() == "Share":
                    # Extract share values
                    share_values = []
                    for col_idx in group_col_indices:
                        cell_val = sheet.cell(row_idx + offset, col_idx).value
                        if isinstance(cell_val, (int, float)):
                            share_values.append(float(cell_val))
                        else:
                            share_values.append(0.0)  # Missing = 0
                    
                    shares_records.append({
                        "ce_item": ce_item,
                        "shares": share_values
                    })
                    break
    
    if not shares_records:
        raise ValueError("No CE item share rows found")
    
    # Convert to DataFrame
    group_ids = _get_group_ids(table_type)
    
    rows = []
    for record in shares_records:
        for group_id, share in zip(group_ids, record['shares']):
            rows.append({
                "ce_item": record["ce_item"],
                "group_id": group_id,
                "raw_share": share
            })
    
    return pd.DataFrame(rows)


def _map_and_aggregate_weights(
    shares_df: pd.DataFrame,
    ce_to_cpi: Dict[str, str],
    excluded_items: set,
    table_type: str
) -> pd.DataFrame:
    """
    Map CE items to CPI categories and aggregate shares into weights.
    
    Handles excluded items and renormalizes to sum=1.0 per group.
    """
    # Mark each row with CPI category or excluded
    shares_df = shares_df.copy()
    shares_df['category_id'] = None
    shares_df['excluded'] = False
    
    for idx, row in shares_df.iterrows():
        ce_item = row['ce_item']
        
        # Try exact match first
        if ce_item in ce_to_cpi:
            shares_df.at[idx, 'category_id'] = ce_to_cpi[ce_item]
        elif ce_item in excluded_items:
            shares_df.at[idx, 'excluded'] = True
        else:
            # Try fuzzy match (case-insensitive, partial)
            matched = False
            for ce_label, cpi_cat in ce_to_cpi.items():
                if ce_label.lower() in ce_item.lower() or ce_item.lower() in ce_label.lower():
                    shares_df.at[idx, 'category_id'] = cpi_cat
                    matched = True
                    break
            
            if not matched:
                # Mark as excluded (unmapped)
                shares_df.at[idx, 'excluded'] = True
    
    # Separate included and excluded
    included_df = shares_df[~shares_df['excluded']].copy()
    excluded_df = shares_df[shares_df['excluded']].copy()
    
    # Aggregate shares by group × category
    aggregated = included_df.groupby(['group_id', 'category_id'])['raw_share'].sum().reset_index()
    aggregated.rename(columns={'raw_share': 'weight'}, inplace=True)
    
    # Convert percent shares to weights (0..1)
    aggregated['weight'] = aggregated['weight'] / 100.0
    
    # Calculate excluded share per group
    excluded_by_group = excluded_df.groupby('group_id')['raw_share'].sum()
    
    # Renormalize weights to sum to 1.0 (excluding excluded items)
    final_weights = []
    
    for group_id in _get_group_ids(table_type):
        group_data = aggregated[aggregated['group_id'] == group_id].copy()
        
        # Compute excluded share for this group
        excluded_share = excluded_by_group.get(group_id, 0.0) / 100.0
        
        # Renormalize included weights
        current_sum = group_data['weight'].sum()
        if current_sum > 0:
            renorm_factor = 1.0 / current_sum
            group_data['weight'] = group_data['weight'] * renorm_factor
        
        # Add excluded_share metadata
        group_data['excluded_share'] = excluded_share
        
        final_weights.append(group_data)
    
    result = pd.concat(final_weights, ignore_index=True)
    
    # Validate weights sum to 1.0 per group
    for group_id in result['group_id'].unique():
        group_sum = result[result['group_id'] == group_id]['weight'].sum()
        if abs(group_sum - 1.0) > 0.001:
            raise ValueError(
                f"Weights for {group_id} sum to {group_sum:.4f}, expected 1.0 ± 0.001"
            )
    
    return result[['group_id', 'category_id', 'weight',  'excluded_share']]


def _get_group_ids(table_type: str) -> List[str]:
    """Get group IDs for quintile or decile."""
    if table_type == "quintile":
        return ["Q1", "Q2", "Q3", "Q4", "Q5"]
    else:  # decile
        return ["D1", "D2", "D3", "D4", "D5", "D6", "D7", "D8", "D9", "D10"]


def save_weights_to_file(
    weights_df: pd.DataFrame,
    output_path: Path,
    weights_year: int,
    table_type: str,
    metadata: Optional[Dict] = None
) -> None:
    """
    Save weights to JSON file matching weights.schema.json format.
    
    Args:
        weights_df: DataFrame with columns [group_id, category_id, weight, excluded_share]
        output_path: Path to save JSON file
        weights_year: CE table year
        table_type: "quintile" or "decile"
        metadata: Optional metadata dict
    """
    # Create output structure
    output = {
        "weights_year": weights_year,
        "grouping": table_type,
        "rows": [],
        "excluded_share": float(weights_df['excluded_share'].iloc[0]) if 'excluded_share' in weights_df.columns else 0.0,
        "metadata": metadata or {}
    }
    
    # Add rows
    for _, row in weights_df.iterrows():
        output["rows"].append({
            "group_id": row['group_id'],
            "category_id": row['category_id'],
            "weight": float(row['weight'])
        })
    
    # Save to file
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, 'w') as f:
        json.dump(output, f, indent=2)
    
    print(f"Saved weights to {output_path}")
    print(f"  Groups: {weights_df['group_id'].nunique()}")
    print(f"  Categories: {weights_df['category_id'].nunique()}")
    print(f"  Total rows: {len(weights_df)}")
